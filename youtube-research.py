import json
import requests
import locale
from openpyxl import Workbook
from openpyxl.worksheet.hyperlink import Hyperlink

locale.setlocale(locale.LC_ALL, 'en_US.UTF-8')

api_key = 'AIzaSyBCnpKiVcn-r543ZoyFaWX8TNi5xihHCiw'


def search_videos(query):
    formatted_keyword = "+".join(query.split(" "))
    url = f"https://www.googleapis.com/youtube/v3/search?q={formatted_keyword}&part=snippet&type=video&maxResults=20&key={api_key}"
    response = requests.get(url)
    if response.ok:
        data = json.loads(response.text)
        if "items" in data:
            video_ids = []
            for item in data["items"]:
                video_ids.append(item['id']['videoId'])
            video_url = f'https://www.googleapis.com/youtube/v3/videos?part=snippet,statistics&id={",".join(video_ids)}&key={api_key}'
            video_response = requests.get(video_url)
            video_data = json.loads(video_response.text)
            videos = []
            for item in video_data["items"]:
                videos.append({
                    "title": item['snippet']['title'],
                    "videoId": f"https://www.youtube.com/watch?v={item['id']}",
                    "viewCount": int(item['statistics']['viewCount'])
                })
            videos = sorted(videos, key=lambda x: x['viewCount'], reverse=True)
            for video in videos:
                view_count = locale.format_string(
                    "%d", video['viewCount'], grouping=True)
                print(
                    f"{video['title']} (Video ID: {video['videoId']}, View Count: {view_count})")
            while True:
                choice = input(
                    "Menu: \n1. Output to excel file \n2. Continue search \n3. Exit \nEnter your choice: ")
                if choice == "1":
                    filename = input("Enter the filename for the excel file: ")
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Search Results"
                    ws.append(["Title", "Video ID", "View Count"])
                    for video in videos:
                        view_count = locale.format_string(
                            "%d", video['viewCount'], grouping=True)
                        ws.append(
                            [video['title'], video['videoId'], view_count])
                        link = Hyperlink(video['videoId'])
                        ws.cell(row=ws.max_row, column=2).hyperlink = link
                        ws.cell(row=ws.max_row, column=3).value = view_count
                    wb.save(filename + ".xlsx")
                    print(f"Search results have been output to {filename}.")

                elif choice == "2":
                    query = input("Enter the keyword: ")
                    search_videos(query)
                elif choice == "3":
                    print("Exiting the program.")
                    exit()
                else:
                    print("Invalid choice. Please enter a valid option.")
        else:
            print("No videos found.")
    else:
        print(f"Error: {response.reason}")


query = input("Enter a keyword to search for: ")
search_videos(query)
